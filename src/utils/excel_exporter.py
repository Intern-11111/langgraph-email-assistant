from __future__ import annotations

import json
import os
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _style_header(ws, row_idx: int = 1):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # blue header
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center


def _append_kv_table(ws, data: Dict[str, Any], title: Optional[str] = None):
    if title:
        ws.append([title])
    ws.append(["Key", "Value"])
    _style_header(ws, ws.max_row)
    for k, v in data.items():
        ws.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
    ws.append([])
    _auto_width(ws)


def _append_list_table(ws, rows: List[Dict[str, Any]], title: Optional[str] = None):
    if not rows:
        if title:
            ws.append([title])
        ws.append(["(no rows)"])
        ws.append([])
        return
    if title:
        ws.append([title])
    headers = list(rows[0].keys())
    ws.append(headers)
    _style_header(ws, ws.max_row)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    ws.append([])
    _auto_width(ws)


def export_outputs_to_excel(
    output_path: str,
    agent_trace: Optional[Dict[str, Any]] = None,
    calendar_output: Optional[Dict[str, Any]] = None,
    contact_output: Optional[Dict[str, Any]] = None,
):
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Section", "Included"])
    _style_header(ws_summary)
    ws_summary.append(["Agent", bool(agent_trace)])
    ws_summary.append(["Calendar", bool(calendar_output)])
    ws_summary.append(["Contact", bool(contact_output)])
    _auto_width(ws_summary)

    # Agent sheets
    if agent_trace:
        ws_agent_final = wb.create_sheet("AgentFinal")
        final = agent_trace.get("final", {})
        _append_kv_table(ws_agent_final, final, title="Final Summary")

        ws_agent_input = wb.create_sheet("AgentInput")
        _append_kv_table(ws_agent_input, agent_trace.get("input", {}), title="Input")

        ws_agent_trace = wb.create_sheet("AgentTrace")
        steps = agent_trace.get("trace", [])
        # Build normalized rows for the trace
        norm_rows: List[Dict[str, Any]] = []
        for s in steps:
            action_input = s.get("action_input", {}) or {}
            obs = s.get("observation", {}) or {}
            tool = None
            if isinstance(action_input, dict):
                tool = action_input.get("tool")
            norm_rows.append({
                "step": s.get("step"),
                "timestamp": s.get("timestamp"),
                "thought": s.get("thought"),
                "action": s.get("action"),
                "tool": tool,
                "action_input": json.dumps(action_input, ensure_ascii=False),
                "observation": json.dumps(obs, ensure_ascii=False),
            })
        _append_list_table(ws_agent_trace, norm_rows, title="Trace Steps")

    # Calendar sheet
    if calendar_output:
        ws_cal = wb.create_sheet("Calendar")
        # Available slots
        slots = calendar_output.get("available_slots", [])
        ws_cal.append(["Available Slots"])
        ws_cal.append(["slot"])
        _style_header(ws_cal, ws_cal.max_row)
        for s in slots:
            ws_cal.append([s])
        ws_cal.append([])
        # Events table
        events = calendar_output.get("events", [])
        if events:
            _append_list_table(ws_cal, events, title="Events")
        else:
            ws_cal.append(["Events"])
            ws_cal.append(["(none)"])
        _auto_width(ws_cal)

    # Contact sheet
    if contact_output:
        ws_contact = wb.create_sheet("Contact")
        _append_kv_table(ws_contact, contact_output, title="Lookup Result")

    # Ensure directory exists
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
