import json
import os
import sys
from agents.react_loop import ReactAgent
from tools.calendar import read_calendar
from tools.contact import lookup_contact
from openpyxl import Workbook


def export_outputs_to_excel(
    output_path: str,
    agent_trace: dict | None = None,
    calendar_output: dict | None = None,
    contact_output: dict | None = None,
):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Section", "Included"])
    ws_summary.append(["Agent", bool(agent_trace)])
    ws_summary.append(["Calendar", bool(calendar_output)])
    ws_summary.append(["Contact", bool(contact_output)])

    if agent_trace:
        ws_final = wb.create_sheet("AgentFinal")
        final = agent_trace.get("final", {}) or {}
        ws_final.append(["Final Summary"]) 
        for k, v in final.items():
            ws_final.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

        ws_input = wb.create_sheet("AgentInput")
        inp = agent_trace.get("input", {}) or {}
        ws_input.append(["Input"]) 
        for k, v in inp.items():
            ws_input.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

        ws_trace = wb.create_sheet("AgentTrace")
        steps = agent_trace.get("trace", []) or []
        ws_trace.append(["step", "timestamp", "thought", "action", "tool", "action_input", "observation"])
        for s in steps:
            action_input = s.get("action_input", {}) or {}
            obs = s.get("observation", {}) or {}
            tool = None
            if isinstance(action_input, dict):
                tool = action_input.get("tool")
            ws_trace.append([
                s.get("step"),
                s.get("timestamp"),
                s.get("thought"),
                s.get("action"),
                tool,
                json.dumps(action_input, ensure_ascii=False),
                json.dumps(obs, ensure_ascii=False),
            ])

    if calendar_output:
        ws_cal = wb.create_sheet("Calendar")
        slots = calendar_output.get("available_slots", []) or []
        ws_cal.append(["Available Slots"]) 
        ws_cal.append(["slot"]) 
        for s in slots:
            ws_cal.append([s])
        ws_cal.append([])
        events = calendar_output.get("events", []) or []
        if events:
            headers = list(events[0].keys())
            ws_cal.append(headers)
            for e in events:
                ws_cal.append([e.get(h) for h in headers])
        else:
            ws_cal.append(["Events"]) 
            ws_cal.append(["(none)"])

    if contact_output:
        ws_contact = wb.create_sheet("Contact")
        ws_contact.append(["Lookup Result"]) 
        for k, v in (contact_output or {}).items():
            ws_contact.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)


def interactive_manage_contacts_and_events(
    contacts_store_path: str | None = None,
    events_store_path: str | None = None,
):

    import datetime

    # Resolve default store paths relative to repo root
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    data_dir = os.path.join(root_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    contacts_path = contacts_store_path or os.path.join(data_dir, "user_contacts.json")
    events_path = events_store_path or os.path.join(data_dir, "user_events.json")

    def _load_json(path: str):
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return [] if path.endswith(".json") else {}
        return []

    def _save_json(path: str, obj):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)

    contacts = _load_json(contacts_path) or []
    events = _load_json(events_path) or []

    if not isinstance(contacts, list):
        contacts = []
    if not isinstance(events, list):
        events = []

    summary = {"added_contacts": 0, "added_events": 0, "lookups": [], "calendar_previewed": False}

    MENU = (
        "\nChoose an option:\n"
        "  1) Add contact\n"
        "  2) Add event\n"
        "  3) Lookup contact (mock + local)\n"
        "  4) Show calendar availability (mock)\n"
        "  5) Save & exit\n"
        "  6) Exit without saving\n"
        "  7) List staged contacts\n"
        "  8) List staged events\n"
    )

    def _parse_dt(prompt: str) -> str | None:
        raw = input(prompt).strip()
        if not raw:
            return None
        # Try several common formats, falling back to ISO if provided
        fmts = [
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M",
            "%m/%d/%Y %H:%M",
        ]
        for fmt in fmts:
            try:
                return datetime.datetime.strptime(raw, fmt).isoformat()
            except ValueError:
                pass
        # Accept already-ISO strings
        try:
            datetime.datetime.fromisoformat(raw)
            return raw
        except Exception:
            print("Could not parse datetime. Skipping time field.")
            return None

    while True:
        print(MENU)
        choice = input("Enter choice [1-6]: ").strip()

        if choice == "1":
            name = input("Name: ").strip()
            email = input("Email: ").strip()
            phone = input("Phone: ").strip()
            if not name and not email:
                print("Contact needs at least a name or email.")
                continue
            contact = {"name": name, "email": email, "phone": phone}
            contacts.append(contact)
            summary["added_contacts"] += 1
            print("Contact staged.")
            print(json.dumps(contact, indent=2))

        elif choice == "2":
            title = input("Event title: ").strip()
            start_iso = _parse_dt("Start (e.g., YYYY-MM-DD HH:MM): ")
            end_iso = _parse_dt("End (optional, same formats): ")
            attendees_raw = input("Attendees emails (comma-separated, optional): ").strip()
            attendees = [a.strip() for a in attendees_raw.split(',') if a.strip()] if attendees_raw else []
            if not title:
                print("Event needs a title.")
                continue
            event = {"title": title, "start": start_iso, "end": end_iso, "attendees": attendees}
            events.append(event)
            summary["added_events"] += 1
            print("Event staged.")

        elif choice == "3":
            q = input("Lookup query (name/email): ").strip()
            if not q:
                print("Empty query.")
                continue
            # 1) Mock lookup via tool
            res = lookup_contact(q)
            # 2) Local lookup over staged + loaded contacts
            ql = q.lower()
            local_found = None
            for c in contacts:
                name = (c.get("name") or "").lower()
                email = (c.get("email") or "").lower()
                if ql in name or ql in email:
                    local_found = c
                    break
            composite = {
                "tool_result": res,
                "local_result": {
                    "found": bool(local_found),
                    "contact": local_found,
                    "source": "staged_and_saved_contacts_in_session",
                },
            }
            summary["lookups"].append(composite)
            print(json.dumps(composite, indent=2))

        elif choice == "4":
            hint = input("Date hint (optional): ").strip() or None
            cal = read_calendar(user_id="me", date_hint=hint)
            summary["calendar_previewed"] = True
            print(json.dumps(cal, indent=2))

        elif choice == "5":
            _save_json(contacts_path, contacts)
            _save_json(events_path, events)
            print(f"Saved contacts to: {contacts_path}")
            print(f"Saved events to: {events_path}")
            break

        elif choice == "6":
            print("Exiting without saving.")
            break

        else:
            print("Invalid choice. Please select 1-6.")

        if choice == "7":
            print("\nStaged contacts (including previously loaded):")
            print(json.dumps(contacts, indent=2))
        elif choice == "8":
            print("\nStaged events (including previously loaded):")
            print(json.dumps(events, indent=2))

    return {
        "contacts_file": contacts_path,
        "events_file": events_path,
        "summary": summary,
    }

    """
    The above code provides an 
    
    Interactive helper to add contacts and events without changing existing flows.

    - Stores contacts in data/user_contacts.json
    - Stores events in data/user_events.json
    - Can preview mock calendar and lookup mock contacts

    Returns a summary dict of actions performed.
    """


def main():
    # Support running without args by using a friendly default
    if len(sys.argv) < 3:
        subject = "Meeting request"
        body = "Can we schedule for tomorrow?"
        print("No arguments provided. Using default subject/body.")
        print("Usage: python src/main.py \"Subject\" \"Body\"")
    else:
        subject = sys.argv[1]
        body = sys.argv[2]

    agent = ReactAgent(max_steps=6)
    trace = agent.run(subject, body, context={"sender": "manager@company.com"})

    print(json.dumps(trace, indent=2))
    print("\nAgent: I've analyzed the email and consulted tools above.")
    try:
        default_out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "agent_outputs.xlsx")
        choice = input("\nAgent: I can export results to Excel. Choose (y=save to default, c=choose path, N=skip) [y/N/c]\nYou: ").strip().lower()
        if choice in ("y", "yes"):
            out_path = default_out
        elif choice in ("c", "choose", "custom"):
            out_path = input(f"Agent: Enter Excel output path [{default_out}]\nYou: ").strip() or default_out
        else:
            out_path = None

        if out_path:
            # Gather tool outputs for inclusion
            cal = read_calendar(user_id="me", date_hint="next available")
            # Try to lookup a useful contact based on sender
            contact = lookup_contact("manager@company.com")
            export_outputs_to_excel(out_path, agent_trace=trace, calendar_output=cal, contact_output=contact)
            print(f"\nAgent: Excel results saved to: {out_path}")
    except Exception as e:
        print("\nAgent: Excel export failed:", e)

    # Optional interactive contact/event management (non-intrusive)
    try:
        manage_choice = input("\nAgent: Would you like me to manage contacts/events now? (y/N)\nYou: ").strip().lower()
        if manage_choice in ("y", "yes"):
            result = interactive_manage_contacts_and_events()
            print("\nAgent: Contacts/Events session summary:")
            print(json.dumps(result, indent=2))
    except Exception as e:
        print("\nAgent: Contacts/events interaction failed:", e)


if __name__ == "__main__":
    main()
